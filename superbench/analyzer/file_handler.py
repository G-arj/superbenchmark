# Copyright (c) Microsoft Corporation.
# Licensed under the MIT license.

"""A module for file related functions in analyzer."""

from pathlib import Path
import re
import json

import jsonlines
import pandas as pd
import yaml
from openpyxl.styles import Alignment
import markdown

from superbench.common.utils import logger


def read_raw_data(raw_data_path):
    """Read raw data from raw_data_path and store them in raw_data_df.

    Args:
        raw_data_path (str): the path of raw data jsonl file

    Returns:
        DataFrame: raw data, node as index, metric name as columns
    """
    p = Path(raw_data_path)
    raw_data_df = pd.DataFrame()
    if not p.is_file():
        logger.error('DataDiagnosis: invalid raw data path - {}'.format(raw_data_path))
        return raw_data_df

    try:
        with p.open(encoding='utf-8') as f:
            for single_node_summary in jsonlines.Reader(f):
                raw_data_df = raw_data_df.append(single_node_summary, ignore_index=True)
        raw_data_df = raw_data_df.rename(raw_data_df['node'])
        raw_data_df = raw_data_df.drop(columns=['node'])
    except Exception as e:
        logger.error('Analyzer: invalid raw data fomat - {}'.format(str(e)))
    return raw_data_df


def read_rules(rule_file=None):
    """Read rule from rule yaml file.

    Args:
        rule_file (str, optional): The path of rule yaml file. Defaults to None.

    Returns:
        dict: dict object read from yaml file
    """
    default_rule_file = Path(__file__).parent / 'rule/default_rule.yaml'
    p = Path(rule_file) if rule_file else default_rule_file
    if not p.is_file():
        logger.error('DataDiagnosis: invalid rule file path - {}'.format(str(p.resolve())))
        return None
    baseline = None
    with p.open() as f:
        baseline = yaml.load(f, Loader=yaml.SafeLoader)
    return baseline


def read_baseline(baseline_file):
    """Read baseline from baseline json file.

    Args:
        baseline_file (str): The path of baseline json file.

    Returns:
        dict: dict object read from json file
    """
    p = Path(baseline_file)
    if not p.is_file():
        logger.error('DataDiagnosis: invalid baseline file path - {}'.format(str(p.resolve())))
        return None
    baseline = None
    with p.open() as f:
        baseline = json.load(f)
    return baseline


def output_excel_raw_data(writer, raw_data_df, sheet_name):
    """Output raw data into 'sheet_name' excel page.

    Args:
        writer (xlsxwriter): xlsxwriter handle
        raw_data_df (DataFrame): the DataFrame to output
        sheet_name (str): sheet name of the excel
    """
    # Output the raw data
    if isinstance(raw_data_df, pd.DataFrame) and not raw_data_df.empty:
        raw_data_df.to_excel(writer, sheet_name, index=True)
    else:
        logger.warning('DataDiagnosis: excel_data_output - {} data_df is empty.'.format(sheet_name))


def output_excel_data_not_accept(writer, data_not_accept_df, rules):
    """Output data_not_accept_df into 'Not Accept' excel page.

    Args:
        writer (xlsxwriter): xlsxwriter handle
        data_not_accept_df (DataFrame): the DataFrame to output
        rules (dict): the rules of DataDiagnosis
    """
    # Get the xlsxwriter workbook objects and init the format
    workbook = writer.book
    color_format_red = workbook.add_format({'bg_color': '#FFC7CE', 'font_color': '#9C0006'})
    percent_format = workbook.add_format({'num_format': '0.00%'})

    # Output the not accept
    if isinstance(data_not_accept_df, pd.DataFrame):
        data_not_accept_df.to_excel(writer, 'Not Accept', index=True)
        if not data_not_accept_df.empty:
            row_start = 1
            row_end = max(row_start, len(data_not_accept_df))
            columns = list(data_not_accept_df.columns)
            worksheet = writer.sheets['Not Accept']

            for rule in rules:
                for metric in rules[rule]['metrics']:
                    col_index = columns.index(metric)
                    # Apply percent format for the columns whose rules are variance type.
                    if rules[rule]['function'] == 'variance':
                        worksheet.conditional_format(
                            row_start,
                            col_index,
                            row_end,
                            col_index,    # start_row, start_col, end_row, end_col
                            {
                                'type': 'no_blanks',
                                'format': percent_format
                            }
                        )
                    # Apply red format if the value violates the rule.
                    if rules[rule]['function'] == 'value' or rules[rule]['function'] == 'variance':
                        match = re.search(r'(>|<|<=|>=|==|!=)(.+)', rules[rule]['criteria'])
                        if not match:
                            continue
                        symbol = match.group(1)
                        condition = float(match.group(2))
                        worksheet.conditional_format(
                            row_start,
                            col_index,
                            row_end,
                            col_index,    # start_row, start_col, end_row, end_col
                            {
                                'type': 'cell',
                                'criteria': symbol,
                                'value': condition,
                                'format': color_format_red
                            }
                        )

        else:
            logger.warning('DataDiagnosis: excel_data_output - data_not_accept_df is empty.')
    else:
        logger.warning('DataDiagnosis: excel_data_output - data_not_accept_df is not DataFrame.')


def output_excel(raw_data_df, data_not_accept_df, output_path, rules):
    """Output the raw_data_df and data_not_accept_df results into excel file.

    Args:
        raw_data_df (DataFrame): raw data
        data_not_accept_df (DataFrame): defective nodes's detailed information
        output_path (str): the path of output excel file
        rules (dict): the rules of DataDiagnosis
    """
    try:
        writer = pd.ExcelWriter(output_path, engine='xlsxwriter')
        # Check whether writer is valiad
        if not isinstance(writer, pd.ExcelWriter):
            logger.error('DataDiagnosis: excel_data_output - invalid file path.')
            return
        output_excel_raw_data(writer, raw_data_df, 'Raw Data')
        output_excel_data_not_accept(writer, data_not_accept_df, rules)
        writer.save()
    except Exception as e:
        logger.error('DataDiagnosis: excel_data_output - {}'.format(str(e)))


def output_json_data_not_accept(data_not_accept_df, output_path):
    """Output data_not_accept_df into jsonl file.

    Args:
        data_not_accept_df (DataFrame): the DataFrame to output
        output_path (str): the path of output jsonl file
    """
    p = Path(output_path)
    try:
        data_not_accept_json = data_not_accept_df.to_json(orient='index')
        data_not_accept = json.loads(data_not_accept_json)
        if not isinstance(data_not_accept_df, pd.DataFrame):
            logger.warning('DataDiagnosis: output json data - data_not_accept_df is not DataFrame.')
            return
        if data_not_accept_df.empty:
            logger.warning('DataDiagnosis: output json data - data_not_accept_df is empty.')
            return
        with p.open('w') as f:
            for node in data_not_accept:
                line = data_not_accept[node]
                line['Index'] = node
                json_str = json.dumps(line)
                f.write(json_str + '\n')
    except Exception as e:
        logger.error('DataDiagnosis: output json data failed, msg: {}'.format(str(e)))


def merge_column_in_excel(ws, row, column):
    """Merge cells in the selected index of column with continuous same contents.

    Args:
        ws (worksheet): the worksheet of the excel to process
        row (int): the max row index to merge
        column (int): the index of the column to merge
    """
    dict_from = {}
    aligncenter = Alignment(horizontal='center', vertical='center')
    # record continuous row index (start, end) with the same content
    for row_index in range(1, row + 1):
        value = str(ws.cell(row_index, column).value)
        if value not in dict_from:
            dict_from[value] = [row_index, row_index]
        else:
            dict_from[value][1] = dict_from[value][1] + 1
    # merge the cells
    for value in dict_from.values():
        if value[0] != value[1]:
            ws.merge_cells(start_row=value[0], start_column=column, end_row=value[1], end_column=column)
    # align center for merged cells
    for i in range(1, row + 1):
        ws.cell(row=i, column=column).alignment = aligncenter


def output_summary_in_excel(raw_data_df, summary, output_path):
    """Output result summary in excel foramt.

    Args:
        raw_data_df (DataFrame): the DataFrame of raw data df
        summary (DataFrame): the DataFrame of summary
        output_path (str): the path of output file
    """
    try:
        writer = pd.ExcelWriter(output_path, engine='openpyxl')
        # check whether writer is valiad
        if not isinstance(writer, pd.ExcelWriter):
            logger.error('ResultSummary: excel_data_output - invalid file path.')
            return
        # output the raw data in 'Raw Data' sheet
        output_excel_raw_data(writer, raw_data_df, 'Raw Data')
        # output the result summary in 'Summary' sheet
        if isinstance(summary, pd.DataFrame) and not summary.empty:
            summary.to_excel(writer, 'Summary', index=False, header=False)
            worksheet = writer.sheets['Summary']
            row = worksheet.max_row
            # merge cells in 'category' column with the same category
            merge_column_in_excel(worksheet, row, 1)
        else:
            logger.error('ResultSummary: excel_data_output - summary is empty.')
        writer.save()
    except Exception as e:
        logger.error('ResultSummary: excel_data_output - {}'.format(str(e)))


def gen_md_table(data, header):
    """Generate table text in markdown format.

    | header[0] | header[1] |
    |     ----  | ----      |
    |     data  | data      |
    |     data  | data      |

    Args:
        data (list): the data in table
        header (list): the header of table

    Returns:
        list: lines of markdown table
    """
    lines = []
    max_width = len(max(data, key=len))
    header[len(header):max_width] = [' ' for i in range(max_width - len(header))]
    align = ['---' for i in range(max_width)]
    lines.append('| {} |\n'.format(' | '.join(header)))
    lines.append('| {} |\n'.format(' | '.join(align)))
    for line in data:
        full_line = [' ' for i in range(max_width)]
        full_line[0:len(line)] = [str(line[i]) for i in range(len(line))]
        lines.append('| {} |\n'.format(' | '.join(full_line)))
    return lines


def gen_md_lines(summary):
    """Generate text in markdown foramt.

    Use category to be the 2nd-header, use tables to show the data

    Args:
        summary (dict): summary dict, the keys are categories, the values are summary lines for the category

    Returns:
        list: lines in markdown format
    """
    lines = []
    for category in summary:
        lines.append('## {}\n'.format(category))
        summary_df = pd.DataFrame(summary[category])
        summary_lines = summary_df.drop(columns=0, axis=1).values.tolist()
        header = ['metric', 'statistics', 'values']
        table_lines = gen_md_table(summary_lines, header)
        lines.extend(table_lines)
        lines.append('\n')
    return lines


def output_summary_in_md(summary, output_path):
    """Output summary in markdown format.

    Args:
        summary (dict): summary dict, the keys are categories, the values are summary lines for the category
        output_path (str): the path of output file
    """
    try:
        lines = gen_md_lines(summary)
        if len(lines) == 0:
            logger.error('ResultSummary: md_data_output failed')
            return
        with open(output_path, 'w') as f:
            f.writelines(lines)
    except Exception as e:
        logger.error('ResultSummary: md_data_output - {}'.format(str(e)))


def output_summary_in_html(summary, output_path):
    """Output summary in html format.

    Args:
        summary (dict): summary dict, the keys are categories, the values are summary lines for the category
        output_path (str): the path of output file
    """
    try:
        lines = gen_md_lines(summary)
        if len(lines) == 0:
            logger.error('ResultSummary: html_data_output failed')
            return
        lines = ''.join(lines)
        html_str = markdown.markdown(lines, extensions=['markdown.extensions.tables'])
        with open(output_path, 'w') as f:
            f.writelines(html_str)
    except Exception as e:
        logger.error('ResultSummary: html_data_output - {}'.format(str(e)))
